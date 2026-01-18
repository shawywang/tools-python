from typing import Set, Dict

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


# pip install openpyxl
# C:\ProgramData\miniconda3\python.exe -m pip install --upgrade pip ; C:\ProgramData\miniconda3\python.exe -m pip install openpyxl

class Handle:
    def __init__(self):
        self.file_path = r"C:\Users\wangxiao\Nutstore\1\我的坚果云\我的文档\入编\26国考\中央机关及其直属机构2026年度考试录用公务员招考简章.xlsx"  # "/Users/wangxiao/Nutstore Files/我的坚果云/我的文档/入编/26国考/"
        self.file_path2 = r"C:\Users\wangxiao\Nutstore\1\我的坚果云\我的文档\入编\26国考\中央机关及其直属机构2026年度考试录用公务员招考简章_filter.xlsx"
        self.file_path3 = r"C:\Users\wangxiao\Nutstore\1\我的坚果云\我的文档\入编\26国考\中央机关及其直属机构2026年度考试录用公务员招考简章_filter2.xlsx"
        self.wb = None

    def del_column(self, sh_name: str, skip_rows: int, inc: Dict[str, Set[str]], excl: Dict[str, Set[str]], excl_eq: Dict[str, Set[str]]):
        ws = self.wb[sh_name]
        rows = list(ws.rows)  # 读成list
        ws.delete_rows(1, ws.max_row)  # 清空sheet

        for k, v in inc.items():
            col_inc_n = column_index_from_string(k) - 1
            row_n = skip_rows
            while row_n < len(rows):
                t = rows[row_n][col_inc_n].value
                if not t:
                    row_n += 1
                    continue
                if not any(s in t for s in v):
                    del rows[row_n]
                else:
                    row_n += 1

        for k, v in excl.items():
            col_excel_n = column_index_from_string(k) - 1
            row_n = skip_rows
            while row_n < len(rows):
                t = rows[row_n][col_excel_n].value
                if not t:
                    row_n += 1
                    continue
                if any(s in t for s in v):
                    del rows[row_n]
                else:
                    row_n += 1

        for k, v in excl_eq.items():
            col_excel_n = column_index_from_string(k) - 1
            row_n = skip_rows
            while row_n < len(rows):
                t = rows[row_n][col_excel_n].value
                if not t:
                    row_n += 1
                    continue
                if any(s == t for s in v):
                    del rows[row_n]
                else:
                    row_n += 1

        for r in rows:
            ws.append([cell.value for cell in r])


def main():
    c = Handle()
    c.wb = load_workbook(c.file_path)
    c.del_column(
        sh_name="中央党群机关", skip_rows=2,
        inc={  # 单元格信息包含如下关键字的，提取
            "M": {"计算机科学与技术"},
            "R": {"无限制", "服务基层项目工作经历"},
        },
        excl={  # 单元格信息包含如下关键字的，剔除
            "W": {"限应届", "限2026", "英语六级", "英语6级", "招考对象为2026", "仅招录应届", "大学生村官", "军队服役", "限女性", "仅限西藏户籍", "仅招录新疆户籍", "限山西省户籍"}
        },
        excl_eq={  # 单元格信息等于如下关键字的，剔除
            "P": {"中共党员"},
            "N": {"仅限硕士研究生", "硕士", "硕士研究生及以上"}
        }
    )

    c.del_column(
        sh_name="中央国家行政机关（本级）", skip_rows=2,
        inc={  # 单元格信息包含如下关键字的，提取
            "M": {"计算机科学与技术"},
            "R": {"无限制", "服务基层项目工作经历"},
        },
        excl={  # 单元格信息包含如下关键字的，剔除
            "W": {"限应届", "限2026", "英语六级", "英语6级", "招考对象为2026", "仅招录应届", "大学生村官", "军队服役", "限女性", "仅限西藏户籍", "仅招录新疆户籍", "限山西省户籍"}
        },
        excl_eq={  # 单元格信息等于如下关键字的，剔除
            "P": {"中共党员"},
            "N": {"仅限硕士研究生", "硕士", "硕士研究生及以上"}
        }
    )

    c.del_column(
        sh_name="中央国家行政机关省级以下直属机构", skip_rows=2,
        inc={  # 单元格信息包含如下关键字的，提取
            "M": {"计算机科学与技术"},
            "R": {"无限制", "服务基层项目工作经历"},
        },
        excl={  # 单元格信息包含如下关键字的，剔除
            "W": {"限应届", "限2026", "英语六级", "英语6级", "招考对象为2026", "仅招录应届", "大学生村官", "军队服役", "限女性", "仅限西藏户籍", "仅招录新疆户籍", "限山西省户籍"}
        },
        excl_eq={  # 单元格信息等于如下关键字的，剔除
            "P": {"中共党员"},
            "N": {"仅限硕士研究生", "硕士", "硕士研究生及以上"}
        }
    )

    c.del_column(
        sh_name="中央国家行政机关参照公务员法管理事业单位", skip_rows=2,
        inc={  # 单元格信息包含如下关键字的，提取
            "M": {"计算机科学与技术"},
            "R": {"无限制", "服务基层项目工作经历"},
        },
        excl={  # 单元格信息包含如下关键字的，剔除
            "W": {"限应届", "限2026", "英语六级", "英语6级", "招考对象为2026", "仅招录应届", "大学生村官", "军队服役", "限女性", "仅限西藏户籍", "仅招录新疆户籍", "限山西省户籍"}
        },
        excl_eq={  # 单元格信息等于如下关键字的，剔除
            "P": {"中共党员"},
            "N": {"仅限硕士研究生", "硕士", "硕士研究生及以上"}
        }
    )
    c.wb.save(c.file_path2)
    c.wb.close()

    # 删除视觉空行：选中内容行末尾第一列的单元格，ctrl+shift+方向键下，右键删除，选择整行，保存。留意到右侧滑块变长了


if __name__ == '__main__':
    main()
